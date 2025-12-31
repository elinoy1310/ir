# exe3/stage4_llm_rag.py
import requests
from pathlib import Path
from sentence_transformers import SentenceTransformer

# ייבוא הפונקציות והקבועים מהקובץ המעודכן (stage3)
from stage3_retrieval import (
    load_chunkpath_to_source,
    load_bm25_store,
    load_dense_store,
    bm25_retrieve,
    dense_retrieve,
    enrich_results,
    uk_count,
    MODEL_NAME,
    change_chanking_method
)

# Ollama settings
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
OLLAMA_MODEL = "llama3.2:1b"

# -------- LLM call --------
def call_ollama(prompt: str) -> str:
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {
            "temperature": 0.15,
            "num_predict": 300
        }
    }
    try:
        r = requests.post(OLLAMA_URL, json=payload, timeout=300)
        r.raise_for_status()
        return r.json().get("response", "").strip()
    except Exception as e:
        return f"Error calling Ollama: {e}"

def build_prompt(query: str, contexts: list[dict]) -> str:
    ctx_lines = []
    for i, c in enumerate(contexts, 1):
        ctx_lines.append(
            f"[{i}] SOURCE_FILE: {c['source_file']}\n"
            f"CHUNK: {c['chunk']}\n"
            f"TEXT: {c['text']}\n"
        )
    print(f"\nBuilt prompt with {len(' '.join(ctx_lines))} context chunks.")
   
    return(
    "You must follow these rules exactly:\n\n"

    "1. Answer the QUESTION using ONLY the CONTEXT below.\n"
    "2. If the CONTEXT does NOT contain the answer, respond with EXACTLY this sentence and NOTHING ELSE:\n"
    "\"I don't know based on the provided context.\"\n"
    "   - In this case, do NOT add explanations, do NOT add sources, do NOT add any other text.\n\n"

    "3. If you DO answer the question:\n"
    "   - Every factual statement MUST include an in-text citation like [1], [2], etc.\n"
    "   - You may ONLY cite source numbers that directly support the statement.\n"
    "   - Do NOT invent or guess sources.\n\n"

    "4. Only IF you answered the question, add a final line in this exact format:\n"
    "   SOURCES: [x], [y]\n\n"

    f"QUESTION: {query}\n\n"
    "CONTEXT:\n"
    + "\n".join(ctx_lines))


def run_rag(query: str, method: str = "hybrid", k: int = 5):
    # שימוש בפונקציות המיובאות
    uk_n = uk_count()
    chunkpath_to_source = load_chunkpath_to_source()
    X_bm25, vocab, bm25_names = load_bm25_store()
    X_emb, dense_names = load_dense_store()
    st_model = SentenceTransformer(MODEL_NAME)

    if method == "bm25":
        retrieved = bm25_retrieve(query, X_bm25, vocab, bm25_names, top_k=k)
    elif method == "dense":
        retrieved = dense_retrieve(query, X_emb, dense_names, st_model, top_k=k)
    # העשרת התוצאות באמצעות הפונקציה מ-stage3
    enriched = enrich_results(retrieved, chunkpath_to_source, max_chars=3000)

    # בניית הפרומפט וקריאה ל-LLM
    prompt = build_prompt(query, enriched)
    answer = call_ollama(prompt)

     # בניית רשימת מקורות
    sources = [
        f"{c['source_file']} ({c['chunk']}) score={c['score']:.4f}"
        for c in enriched
    ]

    # הצגת התוצאה
    print("\n" + "="*90)
    print(f"METHOD={method}  K={k}")
    print("QUESTION:", query)
    print("\nANSWER:\n", answer)
    print("\nTOP CONTEXT SOURCES:")
    for i, c in enumerate(enriched, 1):
        print(f"[{i}] {c['source_file']}  ({c['chunk']})  score={c['score']:.4f}")
    return answer, sources  # החזרת התשובה לשמירה


def run_rag_with_multiple_configs(queries: list, chunk_method: str):
    change_chanking_method(chunk_method)

    for method in ["dense", "bm25"]:
        for k in [3, 5, 8]:
            for i, query in enumerate(queries, 1):

                print(
                    f"RUN QUERY {i} | METHOD={method.upper()} | "
                    f"K={k} | CHUNKING={chunk_method}"
                )

                answer, sources = run_rag(query, method=method, k=k)

                # TXT
                save_answer_to_txt(
                    Path("exe3/outputs/answers.txt"),
                    query, k, method, chunk_method, answer
                )
                save_sources_to_txt(
                    Path("exe3/outputs/sources.txt"),
                    query, k, method, chunk_method, sources
                )

                # EXCEL
                save_answer_to_excel(
                    Path("exe3/outputs/answers.xlsx"),
                    query, k, method, chunk_method, answer
                )
                save_sources_to_excel(
                    Path("exe3/outputs/sources.xlsx"),
                    query, k, method, chunk_method, sources
                )


from pathlib import Path

def save_answer_to_txt(
    filepath: Path,
    query: str,
    k: int,
    method: str,
    chunking: str,
    answer: str
):
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with open(filepath, "a", encoding="utf-8") as f:
        f.write("\n" + "="*120 + "\n")
        f.write(f"QUERY:\n{query}\n\n")
        f.write(f"K = {k}\n")
        f.write(f"{method.upper()} with {chunking} chunking:\n")
        f.write(answer + "\n")

def save_sources_to_txt(
    filepath: Path,
    query: str,
    k: int,
    method: str,
    chunking: str,
    sources: list[str]
):
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with open(filepath, "a", encoding="utf-8") as f:
        f.write("\n" + "="*120 + "\n")
        f.write(f"QUERY:\n{query}\n\n")
        f.write(f"K = {k}\n")
        f.write(f"{method.upper()} with {chunking} chunking – SOURCES:\n")
        for s in sources:
            f.write(f"- {s}\n")


from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

def get_or_create_wb(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def save_answer_to_excel(
    filepath: Path,
    query: str,
    k: int,
    method: str,
    chunking: str,
    answer: str
):
    wb = get_or_create_wb(filepath)
    sheet_name = f"K_{k}"

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append([
            "Query",
            "BM25_fixed",
            "BM25_parent-son",
            "ST_fixed",
            "ST_parent-son"
        ])
    else:
        ws = wb[sheet_name]

    col_map = {
        ("bm25", "fixed"): 2,
        ("bm25", "parent-son"): 3,
        ("dense", "fixed"): 4,
        ("dense", "parent-son"): 5,
    }

    # חיפוש או הוספת שורה לשאילתה
    row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == query:
            row = r
            break

    if row is None:
        ws.append([query, "", "", "", ""])
        row = ws.max_row
    # ✅ עיצוב תא השאילתה (Query)
    query_cell = ws.cell(row, 1)
    query_cell.alignment = Alignment(
        wrapText=True,
        vertical="center",
        horizontal="left"
)
    col = col_map[(method, chunking)]
    cell = ws.cell(row, col)
    cell.value = answer
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    # התאמת רוחב עמודות
    for c in ws.columns:
        ws.column_dimensions[c[0].column_letter].width = 50

    wb.save(filepath)

def save_sources_to_excel(
    filepath: Path,
    query: str,
    k: int,
    method: str,
    chunking: str,
    sources: list[str]
):
    joined_sources = "\n".join(sources)
    save_answer_to_excel(
        filepath,
        query,
        k,
        method,
        chunking,
        joined_sources
    )



if __name__ == "__main__":

    from exe3.run_all_queries import build_queries
    queries = build_queries()
    run_rag_with_multiple_configs(queries, chunk_method="fixed") 
    run_rag_with_multiple_configs(queries, chunk_method="parent-son")
